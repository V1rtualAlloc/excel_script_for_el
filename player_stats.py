# extract clubs info
import requests, re
from openpyxl import Workbook
from bs4 import BeautifulSoup


class Uleb:
    """ initialize season, phase and ULEB competition """
    def __init__(self, competition_website, season, phase):
        self.competition_website = competition_website
        self.season = season
        self.phase = phase
        self.teams = {'team name': [], 'team code': []}
        self.workbook = Workbook()
        self.workbook.active.title = 'Teams'

    def collect_teams(self):
        url_all_teams = 'http://{0}/competition/teams?seasoncode={1}'
        r = requests.get(url_all_teams.format(self.competition_website, self.season))
        soup = BeautifulSoup(r.text, "html.parser")
        l = soup.find('div', class_='teams').find_all('div', class_='RoasterName')
        # find team names and codes
        for item in l:
            anchor_tag = item.a
            self.teams['team name'].append(str(anchor_tag.string))
            team_code_regex = re.search('clubcode=(.*)&', str(anchor_tag['href']))
            team_code = team_code_regex.group(1)
            self.teams['team code'].append(team_code)
        # write to workbook
        self.workbook.active.cell(row=1, column=1).value = 'Teams in season ' + season
        for index, team_name in enumerate(self.teams['team name']):
            self.workbook.active.cell(row=index + 2, column=1).value = team_name

    def get_team_data(self):
        url_team = 'http://{0}/competition/teams/showteam?clubcode={1}&seasoncode={2}'
        for team_code in self.teams['team code']:
            sheet = self.workbook.create_sheet(self.teams['team name'][self.teams['team code'].index(team_code)])
            print(self.teams['team name'][self.teams['team code'].index(team_code)])
            fields = ['Name', 'Position', 'Height', 'Date of Birth', 'Country',
                      'G', 'MIN', 'PT', '2FGM', '2FGA', '3FGM', '3FGA', 'FTM', 'FTA',
                      'ORB', 'DRB', 'TRB', 'ASS', 'STL', 'TO', 'BF', 'BA', 'FC', 'FR', 'PIR']
            for field in fields:
                sheet.cell(row=1, column=fields.index(field) + 1).value = field
            r = requests.get(url_team.format(self.competition_website, team_code, self.season))
            soup = BeautifulSoup(r.text, "html.parser")
            players_info = soup.find_all(class_='item player')
            for player in players_info:
                player_info = player.find(class_='name').a
                player_code_regex = re.search('pcode=(.*)&', str(player_info['href']))
                player_code = player_code_regex.group(1)
                print(player_info.string, player_code)
                self.get_player_data(player_code, sheet)
        self.workbook.save(self.phase + '.xlsx')

    def get_player_data(self, player_code, sheet):
        player_fields = list()
        url_player = 'http://{0}/competition/players/showplayer?pcode={1}&seasoncode={2}#!{3}'
        r = requests.get(url_player.format(self.competition_website, player_code, self.season, self.phase))
        soup = BeautifulSoup(r.text, "html.parser")
        # if the player has no stats whatsoever exit the method
        if not soup.find(id=self.phase):
            return
        player_fields.append(soup.find(class_='player-data').find(class_='name').string)
        player_fields.append(str(soup.find(class_='player-data').find(class_='summary-first').find_all('span').pop().string))
        for item in soup.find(class_='player-data').find(class_='summary-second').find_all('span'):
            player_code_regex = re.search(': (.*)', item.string)
            player_code = player_code_regex.group(1)
            player_fields.append(player_code)
        l = soup.find(id=self.phase).find(class_='TotalFooter').find_all('td')
        del l[1]
        for td in l:
            item = td.find('span')
            if not str(item.string):
                player_fields.append(0)
            elif ':' in str(item.string):
                player_fields.append(int(str(item.string[:item.string.find(':')])))
            elif '/' in str(item.string):
                player_fields.append(int(str(item.string[:item.string.find('/')])))
                player_fields.append(int(str(item.string[item.string.find('/') + 1:])))
            else:
                player_fields.append(int(str(item.string)))
        curent_empty_row = sheet.max_row + 1
        for index, item in enumerate(player_fields):
            sheet.cell(row=curent_empty_row, column=index+1).value = item

euroleague = 'www.euroleague.net'
uleb_cup='www.eurocupbasketball.com'
season = 'U2016'
phase = 'U2016_RS'
obj = Uleb(uleb_cup, season, phase)
obj.collect_teams()
obj.get_team_data()