import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup

# head to head teams data
teams_pace = Workbook()
teams = []
for gamecode in range(1, 264):
    seasoncode = 'E2016'
    url = 'http://www.euroleague.net/main/results/showgame?gamecode={0}&seasoncode={1}'
    r = requests.get(url.format(gamecode, seasoncode))
    soup = BeautifulSoup(r.text, 'html.parser')

    if "Not Found" in soup.text:
        continue
    # name of the host team
    home_team_name = soup.find('div',
                               class_='LocalClubStatsContainer') \
        .find('div',
              class_='eu-team-stats-teamname') \
        .find('span',
              id='ctl00_ctl00_ctl00_ctl00_maincontainer_maincontent_contentpane_boxscorepane_ctl00_LocalClubStats_lblTeamName') \
        .string

    # statistics of the host team
    home_team_totals = []
    for category in soup.find('div', class_='LocalClubStatsContainer') \
            .find('div', class_='TeamStatsMainContainer table-responsive-container') \
            .find('tr', class_='TotalFooter'):
        if category.string is None:
            continue
        if ':' in str(category.string):
            home_team_totals.append(int(category.string[:category.string.find(':')]))
        elif '/' in str(category.string):
            home_team_totals.append(int(category.string[:category.string.find('/')]))
            home_team_totals.append(int(category.string[category.string.find('/') + 1:]))
        elif category.string.isdigit():
            home_team_totals.append(int(category.string))

    # name of the away team
    away_team_name = soup.find('div',
                               class_='RoadClubStatsContainer') \
        .find('div',
              class_='eu-team-stats-teamname') \
        .find('span',
              id='ctl00_ctl00_ctl00_ctl00_maincontainer_maincontent_contentpane_boxscorepane_ctl00_RoadClubStats_lblTeamName') \
        .string

    # statistics of the away team
    away_team_totals = []
    for category in soup.find('div', class_='RoadClubStatsContainer') \
            .find('div', class_='TeamStatsMainContainer table-responsive-container') \
            .find('tr', class_='TotalFooter'):
        if category.string is None:
            continue
        if ':' in str(category.string):
            away_team_totals.append(int(category.string[:category.string.find(':')]))
        elif '/' in str(category.string):
            away_team_totals.append(int(category.string[:category.string.find('/')]))
            away_team_totals.append(int(category.string[category.string.find('/') + 1:]))
        elif category.string.isdigit():
            away_team_totals.append(int(category.string))

    # add both the opponents to lists for away and home teams
    home_team_totals.insert(0, away_team_name)
    away_team_totals.insert(0, home_team_name)

    # adding data for host team to excel worksheet
    if home_team_name not in teams:
        teams.append(home_team_name)
        team_statistics = ['OPP', 'MIN', 'PT', '2FGM', '2FGA', '3FGM', '3FGA', 'FTM', 'FTA', 'ORB', 'DRB',
                           'TRB', 'ASS', 'STL', 'TOV', 'BF', 'BA', 'FC', 'FR', 'PIR']
        teams_pace.create_sheet(home_team_name)
        for stat in team_statistics:
            teams_pace[home_team_name].cell(row=1, column=team_statistics.index(stat) + 1).value = stat

    teams_played = []
    for team in teams_pace[home_team_name]['A']:
        teams_played.append(team.value)

    if away_team_name not in teams_played:
        for index in range(len(home_team_totals)):
            teams_pace[home_team_name].cell(row=len(teams_played) + 1, column=index + 1).value = home_team_totals[index]
    else:
        for index in range(len(home_team_totals)):
            if index > 0:
                teams_pace[home_team_name].cell(row=teams_played.index(away_team_name) + 1, column=index + 1).value += \
                home_team_totals[index]

    # adding data for away team to excel worksheet
    if away_team_name not in teams:
        teams.append(away_team_name)
        team_statistics = ['OPP', 'MIN', 'PT', '2FGM', '2FGA', '3FGM', '3FGA', 'FTM', 'FTA', 'ORB', 'DRB',
                           'TRB', 'ASS', 'STL', 'TOV', 'BF', 'BA', 'FC', 'FR', 'PIR']
        teams_pace.create_sheet(away_team_name)
        for stat in team_statistics:
            teams_pace[away_team_name].cell(row=1, column=team_statistics.index(stat) + 1).value = stat

    teams_played = []
    for team in teams_pace[away_team_name]['A']:
        teams_played.append(team.value)

    if home_team_name not in teams_played:
        for index in range(len(away_team_totals)):
            teams_pace[away_team_name].cell(row=len(teams_played) + 1, column=index + 1).value = away_team_totals[index]
    else:
        for index in range(len(away_team_totals)):
            if index > 0:
                teams_pace[away_team_name].cell(row=teams_played.index(home_team_name) + 1, column=index + 1).value += \
                away_team_totals[index]
    print "Game " + str(gamecode) + " done!"

teams_pace.save('teams_pace.xlsx')
