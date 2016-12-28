# extract clubs info
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup

base_url = 'http://www.euroleague.net'
teams_url = base_url + '/competition/teams'
r = requests.get(teams_url)

soup = BeautifulSoup(r.text, 'html.parser')
tables = soup.find_all('ul', class_='nav-teams nav-teams-16')
# get data for all teams
teams_data = {   
				'teams' : [], 
				'urls' : []
			}
for table in tables:
	lis = table.find_all('li')
	for li in lis:
		teams_data['teams'].append(li.a['title'])
		teams_data['urls'].append(base_url + str(li.a['href']))
	
# create a xlsx file with a sheet 'Teams'
wb = Workbook()
ws = wb.active
ws.title = 'Teams'
ws.cell(row = 1, column = 1).value = 'Team'
for cell in range(len(teams_data['teams'])):
	ws.cell(row = cell + 2, column = 1).value = teams_data['teams'][cell]
	ws.cell(row = cell + 2, column = 1).hyperlink = teams_data['teams'][cell]

# loop all teams and write them in different sheet
for team in teams_data['teams']:
	wb.create_sheet(team)
	player_stats = ['Name', 'Jersey Number', 'Position', 'Country', 'Year of Birth', 'Height',
				   'G', 'GS', 'MIN', 'PT', '2FGM', '2FGA', '3FGM', '3FGA', 'FTM', 'FTA',
				   'ORB', 'DRB', 'TRB', 'ASS', 'STL', 'TO', 'BF', 'BA', 'FC', 'FM', 'PIR']
	# first row is a description about the data for each team
	for item in range(len(player_stats)):
		wb[team].cell(row = 1, column = item + 1).value = player_stats[item]
	
	team_index = teams_data['teams'].index(team)
	r = requests.get(teams_data['urls'][team_index])
	soup = BeautifulSoup(r.text, 'html.parser')
	divs = soup.find_all('div', class_='item player')
	# the actual statistical data for players of a team starts at 2nd row
	row = 2
	print('Processing ' + team),
	for div in divs:
		stat_list = []
		player_name = div.find('div', class_='name').a
		stat_list.append(player_name.string)
		# player url
		url = base_url + player_name['href']
		general_data = div.find('div', class_='data')
		for data in general_data:
			if 'Height' in data.string:
				stat_list.append(data.string[-4:])
			else:
				stat_list.append(data.string)
		# get the statistics for player
		r = requests.get(url)
		soup = BeautifulSoup(r.text, 'html.parser')	
		tds = soup.find_all('td', class_='PlayerTitleColumn')
		for td in tds:
			# here is the magic part, take all of the actual statistical data(FTs, FGs, 2PT, 3TP...) from 'Totals' category
			if 'Totals' in td.string:
				other_stats = td.parent
				for stat in other_stats:
					if stat.string == 'Totals':
						continue
					elif stat.string.find(':') != -1:
						# count only how many minutes player has played, cut the seconds
						stat_list.append(stat.string[:stat.string.find(':')])
					elif stat.string.find('/') != -1:
						# unmerge the categories such as FT made/ FT attended in two columns
						stat_list.append(stat.string[:stat.string.find('/')])
						stat_list.append(stat.string[stat.string.find('/')+1:])
					else:
						# process all other data as it is
						stat_list.append(stat.string)
		# remove the unwanted new line characters from the list
		stat_list = filter(lambda glyph: glyph != u'\n', stat_list)
		for i in range(len(stat_list)):
			wb[team].cell(row = row, column = i + 1, value = stat_list[i])
		wb[team].cell(row = row, column = 1).hyperlink = url	
		print('.'),
		row += 1
	print('Done')
wb.save('data.xlsx')