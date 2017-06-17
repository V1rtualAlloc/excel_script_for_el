# imports from openpyxl library
from openpyxl import Workbook
from openpyxl import load_workbook

# open data workbook to extract the info
data_workbook = load_workbook('data.xlsx')
# generate teams sheet for teams
teams_sheet = tuple(data_workbook["Teams"].columns)
team_names = []
for team in teams_sheet:
    for cell in team:
        if cell.value != 'Team':
            team_names.append(str(cell.value))

# calculate totals of all categories for all players in all teams
for team in team_names:
    current_team_sheet = data_workbook[team]
    players_cols = tuple(current_team_sheet.columns)

    for player in players_cols:
        current_column_value = 0
        for cell in player:
            # 1st row is name of statistical data field (name, position, FGA, REB...)
            # other rows are actual values
            if isinstance(cell.value, (int, long, float, complex)):
                current_column_value += cell.value
        if current_column_value != 0:
            current_team_sheet.cell(row=len(player) + 1, column=players_cols.index(player) + 1,
                                    value=current_column_value)

data_workbook.save('PER.xlsx')
